import openpyxl 
import re
import os
import glob
from itertools import combinations 
from collections import defaultdict 


#Node structure for graph
class Node:

    def __init__(self,src,dest,wt):
        self.src = src
        self.dest = dest
        self.wt = wt


#Class to represent an un-directed graph using adjacency list representation 
class Graph: 
   
    def __init__(self,vertices): 
        self.V = vertices #No. of vertices 
        self.V_org = vertices 
        self.graph = defaultdict(list) # default dictionary to store graph 
        
        
    # function to add an edge to graph 
    def addEdge(self,u,v,w): 
        self.graph[u].append(Node(u,v,w))
        self.graph[v].append(Node(v,u,w))

        
    #function to print graph
    def printGraph(self):
        s = ""
        for i in self.graph:
            s = s + str(i) + " is connected to "
            print(str(i) + " is connected to ")
            for node in self.graph[i]:
                s = s + str(node.dest) + "(Weight = " + str(node.wt) + ")" + " "
                print(str(node.dest) + "(Weight = " + str(node.wt) + ")" + " ")
            s = s + "\n"
            print("\n")
        return s

            
class keywordNodeNetGen:
    
    def __init__(self, path):
        self.path = path
        self.subsetSize = 2
        self.keysPair = []
        self.keywordsList = []
        self.dictKeywords = {}
        self.dictKeyVsPid = {}
        self.idList = []
        self.colName = []
        self.wbObj = openpyxl.load_workbook(self.path)
        self.sheetObj = self.wbObj.active    

    #function to extract list of column headers from .xlsx
    def extractListCol(self):
        maxCol = self.sheetObj.max_column    
        # Loop will print all columns name 
        for i in range(1, maxCol + 1): 
            cellObj = self.sheetObj.cell(row = 1, column = i) 
            self.colName.append(cellObj.value) 
    

    #function to extract column index from its header
    def extractColNumber(self, strIn):
        self.extractListCol()
        for x in self.colName:
            if x == strIn:
                colIndex = self.colName.index(x)+1
        return colIndex


    #function to extract list of keywords in all the research papers
    def extractDictOfKeywords(self):
        maxRow = 1000
        keyIndex = self.extractColNumber('keywords')
        self.dictKeywords = {rowIndex : [keyword.strip() for keyword in re.split(";", self.sheetObj.cell(row = rowIndex, column = keyIndex).value) if keyword] for rowIndex in range(2, maxRow)}
        return self.dictKeywords
        

    #function to create list of keywords in all papers
    def generateKeywordsList(self):
        self.dictKeywords = self.extractDictOfKeywords()
        self.keywordsList = list(set().union(*self.dictKeywords.values()))
        return self.keywordsList
    
    
    #function to find combinations of any 2 keywords
    def rSubset(self): 
        self.keywordsList = self.generateKeywordsList()
        self.keysPair = list(combinations(self.keywordsList, self.subsetSize)) 
        return self.keysPair
    
    
    #function to extract weight
    def pathWeight(self, x, y, dictKeywords, idList):
        edgeWeight = 0
        for idNo in idList:
            if(x in dictKeywords[idNo] and y in dictKeywords[idNo]):
                edgeWeight = edgeWeight + 1
        return edgeWeight
    
    
    #function to generate dictionary having keywords as keys and list of paper ids as value list 
    def generateDictKeyVsPid(self):
        self.keywordsList = self.generateKeywordsList()
        self.idList = self.dictKeywords.keys()
        self.dictKeyVsPid = {key:[pid for pid in self.idList if key in self.dictKeywords[pid]] for key in self.keywordsList}
        return self.dictKeyVsPid

# Driver Function 
if __name__ == "__main__": 
    path = "drive/My Drive/Colab Notebooks/datas.xlsx"
    netGen = keywordNodeNetGen(path)
    dictKeyVsPid = netGen.generateDictKeyVsPid()
    dictKeywords = netGen.extractDictOfKeywords()
    idList = dictKeywords.keys()
    nodesCount = len(netGen.generateKeywordsList())
    keyPairList = netGen.rSubset()
    g = Graph(nodesCount) 
    for i in keyPairList:
        edgeWeight = netGen.pathWeight(i[0],i[1],dictKeywords,idList)
        if edgeWeight>1:
            g.addEdge(i[0], i[1], edgeWeight)
        
        
    g.printGraph()